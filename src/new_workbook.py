#new_workbook
import os 
from openpyxl import Workbook
from openpyxl import load_workbook

def create_new_workbook(fileName):
	""" Returns a workbook instance. Workbook returned will be formatted to mimc the standard timesheet. """

	# Check if file exists 
	pwdPath = os.getcwd()
	pwdPath = pwdPath + "/" + fileName + ".xlsx"
	doesFileExist = os.path.isfile(pwdPath)
	while doesFileExist:
		print ("The file {0} already exists.\nOverwrite file? (y/n): ".format(fileName))
		overWriteFile = input() 
		if overWriteFile == "y":
			doesFileExist = False
		else:
			print ("Enter the new file's name: ")
			fileName = input()
			pwdPath = os.getcwd()
			pwdPath = pwdPath + "/" + fileName + ".xlsx"
			doesFileExist = os.path.isfile(pwdPath)

	# Create a workbook. 
	wb = Workbook()

	# Get worksheet created by default with Workbook().
	ws = wb.active

	# Change default name of worksheet.
	ws.title = "Student Timesheet"

	# Header title.
	ws.header_footer.center_header.text = "\n\n\nDepartment of Computer Science\nStudent Timesheet"
	ws.header_footer.center_header.font_size = 20

	# Merge cells to format student information. 
	ws.merge_cells('A1:C1')
	ws.merge_cells('D2:F2')
	ws.merge_cells('D1:F1')
	ws.merge_cells('A2:C2')

	# Prompt for user information.
	print ("Enter the pay period <mm/dd/yy - mm/dd/yy>")
	payPeriod = input()

	print ("Enter your name:")
	userName = input()

	print ("Enter your UNLV ID:")
	unlvID = input()

	print ("Enter your job title:")
	jobTitle = input()

	ws.cell('A1').value = "Pay Period: {0}".format(payPeriod)
	ws.cell('D1').value = "Name: {0}".format(userName)
	ws.cell('A2').value = "UNLV ID#:{0}".format(unlvID)
	ws.cell('D2').value = "Job Title:{0}".format(jobTitle)

	# Date, time, and hour titles.
	ws.merge_cells('A3:A4')
	ws.cell('A3').value = "DATE"

	ws.merge_cells('B3:E3')
	ws.cell('B3').value = '{:^100}'.format('TIME')

	ws.merge_cells('B4:C4')
	ws.cell('B4').value = '{:^50}'.format('IN')

	ws.merge_cells('D4:E4')
	ws.cell('D4').value = '{:^50}'.format('OUT')

	ws.merge_cells('F3:F4')
	ws.cell('F3').value = '{:^20}'.format('TOTAL') 

	# Blank cells.
	ws.merge_cells('A5:F5')

	# Grand total, student signature, supervisor signature titles. 
	ws.merge_cells('A24:E24')
	ws.cell('A24').value = '{:^200}'.format('GRAND TOTAL')
	ws.merge_cells('A25:F25')
	ws.cell('A25').value = ('Student Signature:')
	ws.merge_cells('A26:F26')
	ws.cell('A26').value = ('Supervisor Signature:')

	ws.print_options.horizontalCentered = True
	ws.print_options.verticalCentered = True

	# Save workbook.
	"""fileExtension = ".xlsx"
	fileName += fileExtension"""	
	wb.save(fileName)
	
	print ("\n\nTimesheet '{0}' was created".format(fileName))

	return wb	
