from openpyxl import Workbook
from openpyxl import load_workbook


	# get name of file to be opened 
	print "Enter the name of the file you want to open"
	fileToOpen = raw_input() 
	fileToOpen += fileExtension
	
	# load workbook from given filename
	wb = load_workbook(filename = fileToOpen)

	# grab current sheet
	ws = wb.active

	''' test write and read
	ws['A7'] = "02-24-16"
	testRead =  ws['A6']
	print testRead.value
	'''

	# save workbook
	wb.save(fileToOpen)
